import sukjabot2_macro, sukjabot2_setting, sukjabot2_define, sukjabot2_function
import discord, datetime, pytz, sys, requests, math

from discord.ext import tasks
from openpyxl import load_workbook



## 설정

bot = sukjabot2_setting.bot
num = 5000



## 봇 메인 명령어

@bot.event
async def on_ready(): # 봇 시작 명령어
    global message_time_start
    global message_time
    message_time_start = datetime.datetime.now()
    message_time = message_time_start
    print("\n----------------------------------------")
    print("SukjaBot Status :: Online")
    print("Bot Name = " + bot.user.name)
    print("Bot ID = " + str(bot.user.id))
    print("Bot Version = " + str(discord.__version__))
    print("Start Time = " + str(message_time_start))
    print("----------------------------------------")
    await bot.change_presence(status=discord.Status.online, activity=discord.Game('마무리'))
    #chat_auto_clear.start()

@bot.event
async def on_member_join(member): # 유저 입장 명령어 
    message_time = datetime.datetime.now()
    embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
    embed.add_field(name=f"공주대 디스코드 서버에 오신것을 환영합니다!",value="왼쪽의 <#756867051657887764> 채널에서 자신의 캠퍼스에 맞는 하트 버튼을 눌러주세요.",inline=False)
    embed.add_field(name=f"심심하면 언제든지 채팅이나 음성채널 들어오셔서 대화해요!",value="원하는 닉네임 있으면 바로 바꿔드려요.",inline=False)
    embed.add_field(name=f"프로필 사진도 그려드립니다!",value="왼쪽의 <#810549255210795029> 채널에서 신청해주세요.",inline=False)
    embed.set_author(name=str(member.name) + " 님이 서버에 입장하셨습니다.",icon_url=member.avatar_url)
    await bot.get_channel(int(sukjabot2_setting.mainch)).send(embed=embed)
    print("[ " + str(message_time) + " ] " + "(공주대) " + str(member.name) + " 님이 서버에 입장하셨습니다.")

@bot.event
async def on_member_remove(member): # 유저 퇴장 명령어 
    embed=discord.Embed(title="", description=f"잘가시게",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
    embed.set_author(name=str(member.name) + " 님이 서버에서 탈주하셨습니다.",icon_url=member.avatar_url)
    await bot.get_channel(int(sukjabot2_setting.mainch)).send(embed=embed)
    print("[ " + str(message_time) + " ] " + "(공주대) " + str(member.name) + " 님이 서버에서 탈주하셨습니다.")

@bot.event
async def on_voice_state_update(member, before, after): # 유저 음성채널 입,퇴장 명령어
    bv_chan = before.channel
    av_chan = after.channel
    if bv_chan == av_chan:
        if member.guild.id == sukjabot2_setting.mainid:
            return
    if bv_chan == None: # 음성채널 들어오기
        if member.guild.id == sukjabot2_setting.mainid:
            embed=discord.Embed(title="", description=f"<@!" + str(member.id) + "> 님이 ``" + str(av_chan) + "`` 채널에 들어왔습니다."\
                ,timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0x5487FE)
            embed.set_author(name=str(member),icon_url=member.avatar_url)
            await bot.get_channel(int(sukjabot2_setting.logch)).send(embed=embed)
    elif av_chan == None: # 음성채널 나가기
        if member.guild.id == sukjabot2_setting.mainid:
            embed=discord.Embed(title="", description=f"<@!" + str(member.id) + "> 님이 ``" + str(bv_chan) + "`` 채널에서 나갔습니다."\
                ,timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
            embed.set_author(name=str(member),icon_url=member.avatar_url)
            await bot.get_channel(int(sukjabot2_setting.logch)).send(embed=embed)
    else: # 음성채널 이동
        if member.guild.id == sukjabot2_setting.mainid:
            embed=discord.Embed(title="", description=f"<@!" + str(member.id) + "> 님이 ``" + str(bv_chan) + "`` 채널에서 ``" + str(av_chan) + "`` 채널으로 이동하였습니다."\
                ,timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFFF75B)
            embed.set_author(name=str(member),icon_url=member.avatar_url)
            await bot.get_channel(int(sukjabot2_setting.logch)).send(embed=embed)

@bot.event
async def on_message_delete(message): # 유저 메세지 삭제 시
    message_time = datetime.datetime.now()
    if message.guild.id == sukjabot2_setting.mainid:
        if message.author.id == sukjabot2_setting.mee:
            print("[ " + str(message_time) + " ] " + "(공주대) 삭제된 메세지 " + str(message.author) + " : 봇 메세지 (MEE6)")
            with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                print(f"[{message_time}] {str(message.author)} : 봇 메세지 (MEE6)", file=text_file)
        elif message.author.id == sukjabot2_setting.ifb:
            print("[ " + str(message_time) + " ] " + "(공주대) 삭제된 메세지 " + str(message.author) + " : 봇 메세지 (IF)")
            with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                print(f"[{message_time}] (삭제된 메세지) {str(message.author)} : 봇 메세지 (IF)", file=text_file)
        else:
            print("[ " + str(message_time) + " ] (공주대) 삭제된 메세지 - [" + str(message.author) + "] : " + message.content )
            with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                print(f"[{message_time}] (삭제된 메세지) {str(message.author)} : {message.content}", file=text_file)
    else:
        print("[ " + str(message_time) + " ] (기타서버) 삭제된 메세지 - [" + str(message.author) + "] : " + message.content )

@bot.event
async def on_message_edit(before, after): # 유저 메세지 수정 시
    message_time = datetime.datetime.now()
    if before.guild.id == sukjabot2_setting.mainid:
        if before.author.id == sukjabot2_setting.mee:
            print("[ " + str(message_time) + " ] (공주대) " + str(before.author) + " : 봇 메세지 (MEE6)")
            with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                print(f"[{message_time}] {str(before.author)} : 봇 메세지 (MEE6)", file=text_file)
        elif before.author.id == sukjabot2_setting.ifb:
            return
        else:
            print("[ " + str(message_time) + " ] (공주대) 수정된 메세지 - [" + str(before.author) + "] : "+ before.content + " → " + after.content)
            with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                print(f"[{message_time}] (수정된 메세지) {str(before.author)} : {before.content} → {after.content}", file=text_file)
    else:
        print("[ " + str(message_time) + " ] (기타서버) 수정된 메세지 - [" + str(before.author) + "] : "+ before.content + " → " + after.content)

@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, discord.ext.commands.errors.CommandNotFound):
        if error == "Command \"ㅈㅂ\" is not found":
            return
        else:
            embed=discord.Embed(title= f"", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xe91e63)
            embed.add_field(name=f"ㅅ가져와 '유저 태그'",value=f"유저 프로필사진 가져오기",inline=False)
            embed.add_field(name=f"ㅅ국어사전 '검색할 단어'",value=f"국어사전 검색 진행 (네이버)",inline=False)
            embed.add_field(name=f"ㅅ영어사전 '검색할 단어'",value=f"영어사전 검색 진행 (네이버)",inline=False)
            embed.add_field(name=f"ㅅ누구",value=f"누가 언제 자신을 태그했는지 확인 가능 (없으면 None 표시)",inline=False)
            embed.add_field(name=f"번역 '번역 전 언어' '번역할 언어' '번역할 내용'",value=f"한국어, 일본어, 영어 번역 진행 중 (예: ㅅ번역 한 일 안녕하세요)",inline=False)
            embed.add_field(name=f"ㅅ코로나",value=f"코로나 확진자 수 출력",inline=False)
            embed.add_field(name=f"ㅅ학식 (캠퍼스) (언제)", value=f"기숙사 식단 정보 출력",inline=False)
            embed.add_field(name=f"ㅅ티아 ㅅ산도 ㅅ태봄 ㅅ콩이 ㅅ겨울 ㅅ쮸",value=f"아무튼 귀여우니까 입력해볼 것",inline=False)
            embed.set_author(name="숙자봇 사용설명서",icon_url=sukjabot2_setting.sjb_avatar_url)
            embed.set_footer(text="문의사항은 @890-KN#2825")
            await ctx.send(embed=embed)

@bot.event
async def on_message(message):
    global chat_status
    global message_time
    message_time = datetime.datetime.now()
    chat_status = False
    global del_user_id
    # 터미널에 메세지 출력
    if not message.guild:
        if message.author.id == sukjabot2_setting.rsj:
            print("[ " + str(message_time) + " ] " + "(개인) " + str(message.author) + " : " + message.content)
            await bot.get_channel(int(sukjabot2_setting.mainch)).send(message.content)
        else:
            print("[ " + str(message_time) + " ] " + "(개인) " + str(message.author) + " : " + message.content)
    else:
        if message.guild.id == sukjabot2_setting.mainid:
            if message.author.id == sukjabot2_setting.mee:
                print("[ " + str(message_time) + " ] " + "(공주대) " + str(message.author) + " : 봇 메세지 (MEE6)")
            elif message.author.id == sukjabot2_setting.ifb:
                with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                    print(f"[{message_time}] (공주대) {str(message.author)} : 봇 메세지 (IF)", file=text_file)
            else:
                if message.author.nick is None:
                    print("[ " + str(message_time) + " ] " + "(공주대) " + str(message.author) + " : " + message.content)
                else:
                    print("[ " + str(message_time) + " ] " + "(공주대) " + str(message.author.nick) + " : " + message.content)
                with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                    print(f"[{message_time}] (공주대) {str(message.author)} : {message.content}", file=text_file)
        elif message.channel.id == sukjabot2_setting.tstch:
                return
        else:
            if message.author.nick is None:
                print("[ " + str(message_time) + " ] " + "(기타서버) " + str(message.author) + " : " + message.content)
            else:
                print("[ " + str(message_time) + " ] " + "(기타서버) " + str(message.author.nick) + " : " + message.content)
            with open("C:/Users/lenya/Documents/DiscordBot/logs/" + message_time.strftime("%Y%m%d") + ".txt", 'a') as text_file:
                print(f"[{message_time}] (기타서버) {str(message.author)} : {message.content}", file=text_file)

    # 채팅창 명령어
    if message.content == "rsts" or message.content == "ㄱㄴㅅㄴ": # 봇 재부팅
        if message.author.id == sukjabot2_setting.rsj or message.author.id == sukjabot2_setting.sjb:
            await message.delete()
            await bot.get_channel(int(sukjabot2_setting.logch)).send("숙자씨 OFF")
            sukjabot2_define.restart_bot()
    elif message.content == "off" or message.content == "OFF": # 봇 종료
        if message.author.id == sukjabot2_setting.rsj or message.author.id == sukjabot2_setting.sjb:
            await message.delete()
            await bot.get_channel(int(sukjabot2_setting.logch)).send("숙자씨 OFF")
            await bot.close()
    elif message.content.startswith("ㅅㅈㅂ") or message.content.startswith("숙자봇"):
        if message.author.id == sukjabot2_setting.rsj or message.author.id == sukjabot2_setting.sjb:
            await message.channel.send("네엡")
    elif message.content.startswith("번역"): # 번역
        transwanttext = message.content.split(" ")
        lan_check = sukjabot2_define.lan_check
        if transwanttext[1] != '한' and transwanttext[1] != '영' and transwanttext[1] != '일':
            await message.channel.send("```번역 기능 사용 방법: 번역 <기존 언어> <번역할 언어> <번역할 내용>\
                \n(예시: 번역 한 영 안녕하세요.)\n\n* 한국어(한), 영어(영), 일본어(일) 번역만 가능합니다.```")
        else:
            try:
                transwanttext[3] != None
            except IndexError:
                await message.channel.send("```경고: 번역할 내용을 입력해주세요.```")
            else:
                Text = ""
                textsize = len(transwanttext)
                textsize = int(textsize)
                for i in range(3, textsize):
                    Text = Text + " " + transwanttext[i]
                while True:
                    if transwanttext[1] == transwanttext[2]:
                        await message.channel.send("```경고: 입력된 두 언어가 동일합니다.```")
                        break
                    elif transwanttext[1] == '한':
                        source_lan = 'ko'
                        target_lan = lan_check(transwanttext[2])
                    elif transwanttext[1] == '영':
                        source_lan = 'en'
                        target_lan = lan_check(transwanttext[2])
                    elif transwanttext[1] == '일':
                        source_lan = 'ja'
                        target_lan = lan_check(transwanttext[2])
                    else:
                        await message.channel.send("```경고: 현재는 한국어, 영어, 일본어 번역만 가능합니다.```")
                        break
                    data = {'text' : Text, 'source' : source_lan, 'target' : target_lan}
                    header = {"X-Naver-Client-Id":sukjabot2_setting.client_id, "X-Naver-Client-Secret":sukjabot2_setting.client_secret}
                    response = requests.post(sukjabot2_setting.url, headers=header, data=data)
                    rescode = response.status_code
                    if(rescode==200):
                        t_data = response.json()
                        embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xC884C8)
                        embed.add_field(name=transwanttext[1],value=Text,inline=False)
                        embed.add_field(name=transwanttext[2],value=t_data['message']['result']['translatedText'],inline=False)
                        embed.set_author(name="숙자봇 한-영-일 번역기",icon_url=sukjabot2_setting.sjb_avatar_url)
                        await message.channel.send(embed=embed)
                        break
                    else:
                        await message.channel.send("```경고: 번역 과정에서 오류가 발생했습니다.```")
    elif message.content == "밀어":
        if message.author.id == sukjabot2_setting.gds or message.author.id == sukjabot2_setting.ajs or message.author.id == sukjabot2_setting.rsj \
            or message.author.id == sukjabot2_setting.sjb or message.author.id == sukjabot2_setting.ycc or message.author.id == sukjabot2_setting.ndm:
            del_user_id = message.author.id
            chat_status = True
            embed=discord.Embed(title="", description=f"ㅅㄱ",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
            embed.set_author(name="채팅창 미는 중..",icon_url=sukjabot2_setting.sjb_avatar_url)
            await message.channel.send(embed=embed)
            await message.channel.purge(limit = int(num))
        else:
            embed=discord.Embed(title="", description=f"구라임 ㅅㄱ",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
            embed.set_author(name="채팅창 미는 중..",icon_url=sukjabot2_setting.sjb_avatar_url)
            await message.channel.send(embed=embed)
    elif message.content == "ㅁㅇ":
        if message.author.id == sukjabot2_setting.rsj:
            chat_status = True
            await message.channel.purge(limit = int(num))
            await bot.get_channel(int(sukjabot2_setting.mainch)).purge(limit = int(num))
    elif message.content.startswith("!clear"):
        del_user_id = message.author.id
    elif message.content == "부검" or message.content == "ㅂㄱ":
        try:
            await message.channel.send("범인은 <@!" + str(del_user_id) + ">")
        except NameError:
            await message.channel.send("범인은 <@!"+ str(message.author.id) + ">")
    elif message.content == "삭제" or message.content == "ㅅㅈ":
        if message.author.id == sukjabot2_setting.rsj:
            await message.delete()
            del(del_user_id)
    elif message.content.startswith("<@"):
        if  message.author.id != sukjabot2_setting.sjb:
            if message.content.startswith("<@!"):
                call_user_1 = message.content
                call_user_2 = call_user_1.split('<@!')[1]
                call_user_3 = call_user_2.split('>')[0] #유저 id 추출
                user_count = message.guild.member_count - 6
                call_user_message = message.content.split(" ")
                call_text = ""
                call_text_size = len(call_user_message)
                call_text_size = int(call_text_size)
                for i in range(1, call_text_size):
                    call_text = call_text + " " + call_user_message[i] #유저 메시지 추출
                # 유저 태그 시 엑셀파일에서 아이디/닉네임 검색 후 기록 (태그한 사람 및 시간)
                for member in message.guild.members:
                    botrole = discord.utils.get(message.guild.roles, name='봇')
                    if botrole not in member.roles:
                        #user_id를 엑셀파일 오픈 후 열 검색
                        load_wb = load_workbook("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx", data_only=True)
                        load_ws = load_wb['User_info']
                        for i in range(2, user_count):
                            if call_user_3 == load_ws.cell(i, 4).value:
                                print("[ " + str(message_time) + " ] " + "메시지 저장됨!")
                                load_ws.cell(i, 5, str(message.author.nick))
                                load_ws.cell(i, 6, str(call_text))
                                load_ws.cell(i, 7, str(message_time.strftime("%Y-%m-%d-%H-%M-%S")))
                                load_wb.save("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx")
                                return
                        return
            else:
                call_user_1 = message.content
                call_user_2 = call_user_1.split('<@')[1]
                call_user_3 = call_user_2.split('>')[0] #유저 id 추출
                user_count = message.guild.member_count - 6
                call_user_message = message.content.split(" ")
                call_text = ""
                call_text_size = len(call_user_message)
                call_text_size = int(call_text_size)
                for i in range(1, call_text_size):
                    call_text = call_text + " " + call_user_message[i] #유저 메시지 추출
                # 유저 태그 시 엑셀파일에서 아이디/닉네임 검색 후 기록 (태그한 사람 및 시간)
                for member in message.guild.members:
                    botrole = discord.utils.get(message.guild.roles, name='봇')
                    if botrole not in member.roles:
                        #user_id를 엑셀파일 오픈 후 열 검색
                        load_wb = load_workbook("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx", data_only=True)
                        load_ws = load_wb['User_info']
                        for i in range(2, user_count):
                            if call_user_3 == load_ws.cell(i, 4).value:
                                print("[ " + str(message_time) + " ] " + "메시지 저장됨!")
                                load_ws.cell(i, 5, str(message.author.nick))
                                load_ws.cell(i, 6, str(call_text))
                                load_ws.cell(i, 7, str(message_time.strftime("%Y-%m-%d-%H-%M-%S")))
                                load_wb.save("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx")
                                return
                        return
    elif message.content == "숙자씨":
        embed=discord.Embed(title= f"", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xC884C8)
        embed.add_field(name=f"ㅅ가져와 '유저 태그'",value=f"유저 프로필사진 가져오기",inline=False)
        embed.add_field(name=f"ㅅ국어사전 '검색할 단어'",value=f"국어사전 검색 진행 (네이버)",inline=False)
        embed.add_field(name=f"ㅅ영어사전 '검색할 단어'",value=f"영어사전 검색 진행 (네이버)",inline=False)
        embed.add_field(name=f"ㅅ누구",value=f"누가 언제 자신을 태그했는지 확인 가능 (없으면 None 표시)",inline=False)
        embed.add_field(name=f"ㅅ번역 '번역 전 언어' '번역할 언어' '번역할 내용'",value=f"한국어, 일본어, 영어 번역 진행 중 (예: ㅅ번역 한 일 안녕하세요)",inline=False)
        embed.add_field(name=f"ㅅ티아 ㅅ산도 ㅅ태봄 ㅅ콩이 ㅅ겨울 ㅅ쮸",value=f"아무튼 귀여우니까 입력해볼 것",inline=False)
        embed.set_author(name="숙자봇 사용설명서",icon_url=sukjabot2_setting.sjb_avatar_url)
        embed.set_footer(text="문의사항은 @890-KN#2825")
        await message.channel.send(embed=embed)
    elif message.content == "기말":
        fe_day, fe_hour, fe_min, fe_sec = sukjabot2_define.fin_exam_strong()
        embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
        embed.add_field(name="조졌어요!",value='기말고사까지 {}일 {}시간 {}분 {}초 남았습니다.'.format(fe_day, fe_hour, fe_min, fe_sec),inline=False)
        embed.set_author(name="모두를 위한 타이머",icon_url=sukjabot2_setting.sjb_avatar_url)
        await message.channel.send(embed=embed)
    elif "깜공군대" in message.content:
        go_hell = datetime.datetime(2021, 8, 31, 14)
        gh_years, gh_days, gh_hours, gh_minutes, gh_seconds = sukjabot2_define.remaining_time(go_hell, message_time)
        embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
        embed.add_field(name="지금을 즐기세요!",value='입대까지 {}일 {}시간 {}분 {}초 남았습니다.'.format(gh_days, gh_hours, gh_minutes, gh_seconds),inline=False)
        embed.set_author(name="깜찍공주만을 위한 타이머",icon_url=sukjabot2_setting.sjb_avatar_url)
        await message.channel.send(embed=embed)
    elif "깜공해방" in message.content:
        go_heaven = datetime.datetime(2023, 3, 2)
        mh_years, mh_days, mh_hours, mh_minutes, mh_seconds = sukjabot2_define.remaining_time(go_heaven, message_time)
        embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
        embed.add_field(name="절대 안옵니다!",value='전역까지 {}년 {}일 {}시간 {}분 {}초 남았습니다.'.format(mh_years, mh_days, mh_hours, mh_minutes, mh_seconds),inline=False)
        embed.set_author(name="깜찍공주만을 위한 타이머",icon_url=sukjabot2_setting.sjb_avatar_url)
        await message.channel.send(embed=embed)
    elif message.content == "종강":
        end_hell = datetime.datetime(2021, 6, 16, 18)
        eh_years, eh_days, eh_hours, eh_minutes, eh_seconds = sukjabot2_define.remaining_time(end_hell, message_time)
        embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
        embed.add_field(name="종강까지",value='{}년 {}일 {}시간 {}분 {}초 남았습니다.'.format(eh_years, eh_days, eh_hours, eh_minutes, eh_seconds),inline=False)
        embed.set_author(name="모두를 위한 종강 타이머",icon_url=sukjabot2_setting.sjb_avatar_url)
        await message.channel.send(embed=embed)
    elif "시발" in message.content:
        await message.delete()
        await message.channel.send("♥이쁜말쓰기♥")

    await bot.process_commands(message)

@bot.command(pass_context=True)
async def 시간(ctx):
    await ctx.message.delete()
    message_time = datetime.datetime.now()
    start_years, start_days, start_hours, start_minutes, start_seconds = sukjabot2_define.remaining_time(message_time, message_time_start)
    embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xC884C8)
    embed.add_field(name="봇 작동",value='{}일 {}시간 {}분 {}초 경과'.format(start_days, start_hours, start_minutes, start_seconds),inline=False)
    embed.set_author(name="봇 작동시간",icon_url=sukjabot2_setting.sjb_avatar_url)
    await ctx.send(embed=embed)

@bot.command(pass_context=True)
async def testmain(ctx):
    if ctx.author.id == sukjabot2_setting.rsj:
        return

# 채팅 자동밀기
'''@tasks.loop(seconds=60)
async def chat_auto_clear():
    current_time = datetime.datetime.now() - message_time
    ct_days, ct_hours, ct_minutes, ct_seconds = sukjabot2_define.time_division(current_time)
    if chat_status == True:
        if ct_minutes > 15:
            embed=discord.Embed(title="", description=f"15분 경과 ㅅㄱ",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
            embed.set_author(name="채팅창 자동 삭제 중..",icon_url=sjb_avatar_url)
            await bot.get_channel(int(mainch)).send(embed=embed)
            await bot.get_channel(int(mainch)).purge(limit = int(num))
    else:
        return'''

    
bot.run(sukjabot2_setting.token)