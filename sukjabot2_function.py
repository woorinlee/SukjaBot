import sukjabot2_setting, sukjabot2_define
import discord, datetime, pytz, requests, re, random, time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Color
from bs4 import BeautifulSoup

## 설정

bot = sukjabot2_setting.bot
sjb_avatar_url = sukjabot2_setting.sjb_avatar_url


## 명령어

# 주요 기능
@bot.command(pass_context=True)
async def 코로나(ctx):
    dome_num, over_num, plus_num, time_data = sukjabot2_define.corona_message()
    embed=discord.Embed(title="COVID-19 현황표", description=time_data,timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xF14952)
    embed.add_field(name="국내 발생",value=dome_num + " 명",inline=True)
    embed.add_field(name="해외 유입",value=over_num + " 명",inline=True)
    embed.add_field(name="총",value=str(plus_num) + " 명",inline=True)
    embed.add_field(name="관련 정보 더보기",value="http://ncov.mohw.go.kr/",inline=True)
    embed.set_author(name="숙자봇 코로나 알리미",icon_url=sjb_avatar_url)
    await ctx.send(embed=embed)

@bot.command()
async def 영어사전(ctx, word):
    await ctx.message.delete()
    url = "https://dict.naver.com/search.nhn?query=" + word
    webpage = requests.get(url)
    soup = BeautifulSoup(webpage.content, "html.parser")
    embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xC884C8)
    embed.add_field(name=word+" 검색을 진행합니다.",value=soup.dd.get_text(),inline=False)
    embed.set_author(name="네이버 영어사전",icon_url=sjb_avatar_url)
    await ctx.send(embed=embed)

@bot.command()
async def 국어사전(ctx, word):
    await ctx.message.delete()
    url = "https://dict.naver.com/search.nhn?dicQuery=" + word + "&query=" + word + "&target=dic&ie=utf8&query_utf=&isOnlyViewEE="
    webpage = requests.get(url)
    soup = BeautifulSoup(webpage.content, "html.parser")
    title = soup.select_one('#content > div.kr_dic_section.search_result.dic_kr_entry > ul > li:nth-child(1) > p:nth-child(2)')
    title_text = title.get_text()
    title_text = title_text.replace("\r\n", "")
    title_text = title_text.replace("\t", "")
    result = re.sub('[0~9]+','test',title_text)
    embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xC884C8)
    embed.add_field(name=word+" 검색을 진행합니다.",value=result,inline=False)
    embed.set_author(name="네이버 국어사전",icon_url=sjb_avatar_url)
    await ctx.send(embed=embed)



# 학식 관련 기능
@bot.command(pass_context=True)
async def 학식(ctx, campus, when):
    today = datetime.datetime.now().strftime("%m월 %d일")
    if campus == "예산":
        y_url = "https://www.kongju.ac.kr/kor/38/foods/CAT111"
        y_webpage = requests.get(y_url)
        y_soup = BeautifulSoup(y_webpage.content, "html.parser")
        y_mdatelist = []
        y_mdowlist = []
        y_mlmeallist = []
        y_mdmeallist = []
        # 일자
        for i in range(0, 7):
            j = i + 2
            y_mdate = y_soup.select_one('#restaurants > div.table_scroll > div > table > thead > tr:nth-child(1) > th:nth-child(' + str(j) + ')')
            y_mdatelist.append(sukjabot2_define.get_yesan_meal_date_text(y_mdate))
        #요일
        for i in range(0, 7):
            j = i + 1
            y_mdow = y_soup.select_one('#restaurants > div.table_scroll > div > table > thead > tr:nth-child(2) > th:nth-child(' + str(j) + ')')
            y_mdowlist.append(sukjabot2_define.dow_eng2kor(sukjabot2_define.get_yesan_meal_date_text(y_mdow)))
        #중식
        for i in range(0, 7):
            j = i + 2
            y_mlmeal = y_soup.select_one('#restaurants > div.table_scroll > div > table > tbody > tr:nth-child(1) > td:nth-child(' + str(j) + ')')
            y_mlmeallist.append(sukjabot2_define.get_yesan_meal_text(y_mlmeal))
        #석식
        for i in range(0, 7):
            j = i + 2
            y_mdmeal = y_soup.select_one('#restaurants > div.table_scroll > div > table > tbody > tr:nth-child(2) > td:nth-child(' + str(j) + ')')
            y_mdmeallist.append(sukjabot2_define.get_yesan_meal_text(y_mdmeal))

        if when == "이번주":
            embed=discord.Embed(title="예산캠퍼스 기숙사 중식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=y_mdatelist[i] + " (" + y_mdowlist[i] + ")",value=y_mlmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
            embed=discord.Embed(title="예산캠퍼스 기숙사 석식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=y_mdatelist[i] + " (" + y_mdowlist[i] + ")",value=y_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "오늘":
            embed=discord.Embed(title="예산캠퍼스 기숙사 오늘자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                if y_mdatelist[i] == today:
                    embed.add_field(name=y_mdatelist[i] + " (" + y_mdowlist[i] + ")\n중식",value=y_mlmeallist[i],inline=True)
                    embed.add_field(name="ㅤ\n석식",value=y_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "내일":
            embed=discord.Embed(title="예산캠퍼스 기숙사 내일자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            try:
                for i in range(0, 7):
                    if y_mdatelist[i] == today:
                        k = i + 1
                        embed.add_field(name=y_mdatelist[k] + " (" + y_mdowlist[k] + ")\n중식",value=y_mlmeallist[k],inline=True)
                        embed.add_field(name="ㅤ\n석식",value=y_mdmeallist[k],inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
            except IndexError:
                embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
                embed.add_field(name="학식 알리미 도움말", value="내일자 학식 정보가 없습니다.\n정보 업데이트를 기다려 주세요.",inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
        else:
            embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            embed.add_field(name="학식 알리미 도움말", value="ㅅ학식 예산 오늘 : 오늘자 학식 정보 출력\nㅅ학식 예산 내일 : 내일자 학식 정보 출력\
                \nㅅ학식 예산 이번주 : 이번주 학식 정보 출력",inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
    elif campus == "드림":
        d_url = "https://dormi.kongju.ac.kr/HOME/sub.php?code=041302"
        d_webpage = requests.get(d_url)
        d_soup = BeautifulSoup(d_webpage.content, "html.parser")
        d_mdatelist = []
        d_mdowlist = []
        d_mbmeallist = []
        d_mlmeallist = []
        d_mdmeallist = []
        # 일자
        for i in range(0, 7):
            j = i + 1
            d_mdate = d_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(2)')
            d_mdatelist.append(sukjabot2_define.get_domi_meal_date_text(d_mdate))
        #요일
        for i in range(0, 7):
            j = i + 1
            d_mdow = d_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td.noedge-l.first > font')
            d_mdowlist.append(sukjabot2_define.get_domi_meal_date_text(d_mdow))
        #조식
        for i in range(0, 7):
            j = i + 1
            d_mbmeal = d_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(3)')
            d_mbmeallist.append(sukjabot2_define.get_dream_domi_meal_text(d_mbmeal))
        #중식
        for i in range(0, 7):
            j = i + 1
            d_mlmeal = d_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(4)')
            d_mlmeallist.append(sukjabot2_define.get_dream_domi_meal_text(d_mlmeal))
        #석식
        for i in range(0, 7):
            j = i + 1
            d_mdmeal = d_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td.noedge-r.last')
            d_mdmeallist.append(sukjabot2_define.get_dream_domi_meal_text(d_mdmeal))

        if when == "이번주":
            embed=discord.Embed(title="신관캠퍼스 드림하우스 기숙사 조식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=d_mdatelist[i] + " (" + d_mdowlist[i] + ")",value=d_mbmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
            embed=discord.Embed(title="신관캠퍼스 드림하우스 기숙사 중식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=d_mdatelist[i] + " (" + d_mdowlist[i] + ")",value=d_mlmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
            embed=discord.Embed(title="신관캠퍼스 드림하우스 기숙사 석식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=d_mdatelist[i] + " (" + d_mdowlist[i] + ")",value=d_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "오늘":
            embed=discord.Embed(title="신관캠퍼스 드림하우스 기숙사 오늘자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                if d_mdatelist[i] == today:
                    embed.add_field(name=d_mdatelist[i] + " (" + d_mdowlist[i] + ")\n조식",value=d_mbmeallist[i],inline=True)
                    embed.add_field(name="ㅤ\n중식",value=d_mlmeallist[i],inline=True)
                    embed.add_field(name="ㅤ\n석식",value=d_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "내일":
            embed=discord.Embed(title="신관캠퍼스 드림하우스 기숙사 내일자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            try:
                for i in range(0, 7):
                    if d_mdatelist[i] == today:
                        k = i + 1
                        embed.add_field(name=d_mdatelist[k] + " (" + d_mdowlist[k] + ")\n조식",value=d_mbmeallist[k],inline=True)
                        embed.add_field(name="ㅤ\n중식",value=d_mlmeallist[k],inline=True)
                        embed.add_field(name="ㅤ\n석식",value=d_mdmeallist[k],inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
            except IndexError:
                embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
                embed.add_field(name="학식 알리미 도움말", value="내일자 학식 정보가 없습니다.\n정보 업데이트를 기다려 주세요.",inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
        else:
            embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            embed.add_field(name="학식 알리미 도움말", value="ㅅ학식 드림 오늘 : 오늘자 학식 정보 출력\nㅅ학식 드림 내일 : 내일자 학식 정보 출력\
                \nㅅ학식 드림 이번주 : 이번주 학식 정보 출력",inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
    elif campus == "은행사" or campus == "비전":
        s_url = "https://dormi.kongju.ac.kr/HOME/sub.php?code=041301"
        s_webpage = requests.get(s_url)
        s_soup = BeautifulSoup(s_webpage.content, "html.parser")
        s_mdatelist = []
        s_mdowlist = []
        s_mbmeallist = []
        s_mlmeallist = []
        s_mdmeallist = []
        # 일자
        for i in range(0, 7):
            j = i + 1
            s_mdate = s_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(2)')
            s_mdatelist.append(sukjabot2_define.get_domi_meal_date_text(s_mdate))
        #요일
        for i in range(0, 7):
            j = i + 1
            s_mdow = s_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td.noedge-l.first > font')
            s_mdowlist.append(sukjabot2_define.get_domi_meal_date_text(s_mdow))
        #조식
        for i in range(0, 7):
            j = i + 1
            s_mbmeal = s_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(3)')
            s_mbmeallist.append(sukjabot2_define.get_dream_domi_meal_text(s_mbmeal))
        #중식
        for i in range(0, 7):
            j = i + 1
            s_mlmeal = s_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(4)')
            s_mlmeallist.append(sukjabot2_define.get_dream_domi_meal_text(s_mlmeal))
        #석식
        for i in range(0, 7):
            j = i + 1
            s_mdmeal = s_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td.noedge-r.last')
            s_mdmeallist.append(sukjabot2_define.get_dream_domi_meal_text(s_mdmeal))

        if when == "이번주":
            embed=discord.Embed(title="신관캠퍼스 은행사/비전하우스 기숙사 조식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=s_mdatelist[i] + " (" + s_mdowlist[i] + ")",value=s_mbmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
            embed=discord.Embed(title="신관캠퍼스 은행사/비전하우스 기숙사 중식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=s_mdatelist[i] + " (" + s_mdowlist[i] + ")",value=s_mlmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
            embed=discord.Embed(title="신관캠퍼스 은행사/비전하우스 기숙사 석식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=s_mdatelist[i] + " (" + s_mdowlist[i] + ")",value=s_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "오늘":
            embed=discord.Embed(title="신관캠퍼스 은행사/비전하우스 기숙사 오늘자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                if s_mdatelist[i] == today:
                    embed.add_field(name=s_mdatelist[i] + " (" + s_mdowlist[i] + ")\n조식",value=s_mbmeallist[i],inline=True)
                    embed.add_field(name="ㅤ\n중식",value=s_mlmeallist[i],inline=True)
                    embed.add_field(name="ㅤ\n석식",value=s_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "내일":
            embed=discord.Embed(title="신관캠퍼스 은행사/비전하우스 기숙사 내일자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            try:
                for i in range(0, 7):
                    if s_mdatelist[i] == today:
                        k = i + 1
                        embed.add_field(name=s_mdatelist[k] + " (" + s_mdowlist[k] + ")\n조식",value=s_mbmeallist[k],inline=True)
                        embed.add_field(name="ㅤ\n중식",value=s_mlmeallist[k],inline=True)
                        embed.add_field(name="ㅤ\n석식",value=s_mdmeallist[k],inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
            except IndexError:
                embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
                embed.add_field(name="학식 알리미 도움말", value="내일자 학식 정보가 없습니다.\n정보 업데이트를 기다려 주세요.",inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
        else:
            embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            embed.add_field(name="학식 알리미 도움말", value="ㅅ학식 은행사 오늘 (or ㅅ학식 비전 오늘) : 오늘자 학식 정보 출력\
                \nㅅ학식 은행사 내일 (or ㅅ학식 비전 내일) : 내일자 학식 정보 출력\nㅅ학식 은행사 이번주 (or ㅅ학식 비전 이번주) : 이번주 학식 정보 출력",inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
    elif campus == "천안":
        c_url = "https://dormi.kongju.ac.kr/sub.php?code=041303"
        c_webpage = requests.get(c_url)
        c_soup = BeautifulSoup(c_webpage.content, "html.parser")
        c_mdatelist = []
        c_mdowlist = []
        c_mlmeallist = []
        c_mdmeallist = []
        # 일자
        for i in range(0, 7):
            j = i + 1
            c_mdate = c_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(2)')
            c_mdatelist.append(sukjabot2_define.get_domi_meal_date_text(c_mdate))
        #요일
        for i in range(0, 7):
            j = i + 1
            c_mdow = c_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td.noedge-l.first > font')
            c_mdowlist.append(sukjabot2_define.get_domi_meal_date_text(c_mdow))
        #중식
        for i in range(0, 7):
            j = i + 1
            c_mlmeal = c_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td:nth-child(4)')
            c_mlmeallist.append(sukjabot2_define.get_cheonan_domi_meal_text(c_mlmeal))
        #석식
        for i in range(0, 7):
            j = i + 1
            c_mdmeal = c_soup.select_one('#food-info > div > table > tbody > tr:nth-child(' + str(j) + ') > td.noedge-r.last')
            c_mdmeallist.append(sukjabot2_define.get_cheonan_domi_meal_text(c_mdmeal))
        if when == "이번주":
            embed=discord.Embed(title="천안캠퍼스 챌린지하우스 기숙사 중식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=c_mdatelist[i] + " (" + c_mdowlist[i] + ")",value=c_mlmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
            embed=discord.Embed(title="천안캠퍼스 챌린지하우스 기숙사 석식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                embed.add_field(name=c_mdatelist[i] + " (" + c_mdowlist[i] + ")",value=c_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "오늘":
            embed=discord.Embed(title="천안캠퍼스 챌린지하우스 기숙사 오늘자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            for i in range(0, 7):
                if c_mdatelist[i] == today:
                    embed.add_field(name=c_mdatelist[i] + " (" + c_mdowlist[i] + ")\n중식",value=c_mlmeallist[i],inline=True)
                    embed.add_field(name="ㅤ\n석식",value=c_mdmeallist[i],inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
        elif when == "내일":
            embed=discord.Embed(title="천안캠퍼스 챌린지하우스 기숙사 내일자 학식", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            try:
                for i in range(0, 7):
                    if c_mdatelist[i] == today:
                        k = i + 1
                        embed.add_field(name=c_mdatelist[k] + " (" + c_mdowlist[k] + ")\n중식",value=c_mlmeallist[k],inline=True)
                        embed.add_field(name="ㅤ\n석식",value=c_mdmeallist[k],inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
            except IndexError:
                embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
                embed.add_field(name="학식 알리미 도움말", value="내일자 학식 정보가 없습니다.\n정보 업데이트를 기다려 주세요.",inline=True)
                embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
                await ctx.send(embed=embed)
        else:
            embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
            embed.add_field(name="학식 알리미 도움말", value="ㅅ학식 천안 오늘 : 오늘자 학식 정보 출력\nㅅ학식 천안 내일 : 내일자 학식 정보 출력\
                \nㅅ학식 천안 이번주 : 이번주 학식 정보 출력",inline=True)
            embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
            await ctx.send(embed=embed)
    else:
        embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFF9D5B)
        embed.add_field(name="학식 알리미 도움말", value="ㅅ학식 천안 오늘/내일/이번주 : 천안캠퍼스 챌린지하우스 학식 정보 출력\
            \nㅅ학식 드림 오늘/내일/이번주 : 신관캠퍼스 드림하우스 학식 정보 출력\nㅅ학식 은행사/비전 오늘/내일/이번주 : 신관캠퍼스 은행사/비전하우스 학식 정보 출력\
            \nㅅ학식 예산 오늘/내일/이번주 : 예산캠퍼스 금오사/예지사 학식 정보 출력",inline=True)
        embed.set_author(name="숙자봇 학식 알리미",icon_url=sjb_avatar_url)
        await ctx.send(embed=embed)


# 유저 관련 기능
@bot.command(pass_context=True)
async def 누구(ctx):
    role = discord.utils.get(ctx.guild.roles, name='봇')
    for member in ctx.guild.members:
        if role not in member.roles:
            user_id = ctx.author.id
            user_count = ctx.guild.member_count - 6
            load_wb = load_workbook("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx", data_only=True)
            load_ws = load_wb['User_info']
            for i in range(2, user_count):
                if str(user_id) == str(load_ws.cell(i, 4).value):
                    call_message = "내용 없음"
                    call_user = load_ws.cell(i, 5).value
                    call_message = load_ws.cell(i, 6).value
                    call_time = load_ws.cell(i, 7).value
                    
                    embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFFF75B)
                    embed.add_field(name=f"언급한 유저",value=call_user,inline=False)
                    embed.add_field(name=f"언급 메세지 내용",value=call_message,inline=False)
                    embed.add_field(name=f"언급한 시간",value=call_time,inline=False)
                    embed.set_author(name=str(ctx.author.nick)+" 님을 찾는 유저입니다.",icon_url=ctx.author.avatar_url)
                    await ctx.send(embed=embed)
                    return
            return

@bot.command()
async def 가져와(ctx, member: discord.Member):
    await ctx.send(member.avatar_url)

@bot.command(pass_context=True)
async def 멤버(ctx):
    if ctx.author.id == sukjabot2_setting.rsj:
        await ctx.message.delete()
        write_wb = Workbook()
        write_ws = write_wb.create_sheet('User_info')
        role = discord.utils.get(ctx.guild.roles, name='봇')
        i = 2
        for member in ctx.guild.members:
            if role not in member.roles:
                if member.nick is None:
                    write_ws.cell(i, 1, i-1)
                    write_ws.cell(i, 3, member.name)
                    write_ws.cell(i, 4, str(member.id))
                    i += 1
                else:
                    write_ws.cell(i, 1, i-1)
                    write_ws.cell(i, 2, member.nick)
                    write_ws.cell(i, 3, member.name)
                    write_ws.cell(i, 4, str(member.id))
                    i += 1
                write_wb.save("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx")

@bot.command(pass_context=True)
async def 캠퍼스(ctx):
    cnum = 0
    knum = 0
    ynum = 0
    user_count = ctx.guild.member_count - 6
    load_wb = load_workbook("C:/Users/lenya/Documents/DiscordBot/DiscordServer_Data.xlsx", data_only=True)
    load_ws = load_wb['User_info']
    for i in range(2, user_count+2):
        user_camp = load_ws.cell(i, 2).value
        if "천안" in user_camp:
            cnum += 1
        elif "신관" in user_camp:
            knum += 1
        elif "예산" in user_camp:
            ynum += 1
        else:
            cnum += 1
    embed=discord.Embed(title="", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xFFF75B)
    embed.add_field(name="현황",value='천안 : {}명 / 신관 : {}명 / 예산 : {}명'.format(cnum, knum, ynum),inline=False)
    embed.add_field(name="총원",value='현재 서버 총 인원은 {}명 입니다.'.format(user_count),inline=False)
    embed.set_author(name="디스코드 서버 인원 정보",icon_url=sjb_avatar_url)
    await ctx.send(embed=embed)

@bot.command()
async def 도움(ctx):
    embed=discord.Embed(title= f"", description=f"",timestamp=datetime.datetime.now(pytz.timezone('UTC')),color=0xe91e63)
    embed.add_field(name=f"ㅅ가져와 '유저 태그'",value=f"유저 프로필사진 가져오기",inline=False)
    embed.add_field(name=f"ㅅ국어사전 '검색할 단어'",value=f"국어사전 검색 진행 (네이버)",inline=False)
    embed.add_field(name=f"ㅅ영어사전 '검색할 단어'",value=f"영어사전 검색 진행 (네이버)",inline=False)
    embed.add_field(name=f"ㅅ누구",value=f"누가 언제 자신을 태그했는지 확인 가능 (없으면 None 표시)",inline=False)
    embed.add_field(name=f"번역 '번역 전 언어' '번역할 언어' '번역할 내용'",value=f"한국어, 일본어, 영어 번역 진행 중 (예: ㅅ번역 한 일 안녕하세요)",inline=False)
    embed.add_field(name=f"ㅅ코로나",value=f"코로나 확진자 수 출력",inline=False)
    embed.add_field(name=f"ㅅ티아 ㅅ산도 ㅅ태봄 ㅅ콩이 ㅅ겨울 ㅅ쮸",value=f"아무튼 귀여우니까 입력해볼 것",inline=False)
    embed.set_author(name="숙자봇 사용설명서",icon_url=sjb_avatar_url)
    embed.set_footer(text="문의사항은 @890-KN#2825")
    await ctx.send(embed=embed)

@bot.command()
async def 비상(ctx, num1):
    await ctx.message.delete()
    num3 = int(num1)
    if num3 < 6:
        for num1 in range(0, num3):
            await ctx.send("<:bi:771691070403772456> <:sang:771691082366058496>")
    else:
        await ctx.send("```경고: 너무 많은 수를 입력했습니다. (5 이하 입력)```")



# 봇 테스트용
@bot.command(pass_context=True)
async def 낚시(ctx):
    if ctx.author.id == sukjabot2_setting.rsj:
        wait_message = ["큰 물고기 낚였으면 좋겠다...", "아직 찌를 물지 않은 듯하다...", "해물탕 먹고 싶다...", "하늘이 참 맑다...", ""]
        fail_message = ["찌를 올렸지만 아무 것도 없었다...", "자리를 잘못 잡았나...?", "물고기가 떠나가 버렸다..."]
        fishing_user_memtion = "<@!" + str(ctx.author.id) + ">"
        main_embed=discord.Embed(title= f":sweat_drops:  낚시찌를 던졌다! (첨벙)", description=f"```cs\n※ 느낌이 오면 🎣를 '연타'하자!\n(그만하려면 🚫을 누르자)```",color=0x822659)
        fishing_msg = await ctx.send(fishing_user_memtion, embed=main_embed)
        await fishing_msg.add_reaction("🎣")
        await fishing_msg.add_reaction("🚫")
        time.sleep(2.5)
        for i in range(1, 6):
            wait_embed=discord.Embed(title= f"기다리는 중...", description=wait_message[random.randint(0,3)] + "```cs\n※ 느낌이 오면 🎣를 '연타'하자!\n(그만하려면 🚫을 누르자)```",color=0x822659)
            await fishing_msg.edit(embed=wait_embed)
            time.sleep(2.5)
        fail_embed=discord.Embed(title= f"낚시 실패", description=fail_message[random.randint(0,2)],color=0x979C9F)
        await fishing_msg.edit(embed=fail_embed)
        

@bot.command()
async def 핑(ctx):
    ping_ = bot.latency
    ping = round(ping_ * 1000)
    await ctx.send(f"{ping}ms 입니다.")


# 사진 출력
@bot.command(pass_context=True)
async def 산도(ctx):
    rannum = random.randrange(1, 72)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Sando/sd (" + str(rannum) + ").jpg"))

@bot.command(pass_context=True)
async def 티아(ctx):
    rannum = random.randrange(1, 28)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Tia/ta (" + str(rannum) + ").jpg"))

@bot.command(pass_context=True)
async def 콩이(ctx):
    rannum = random.randrange(1, 34)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Kongi/ki (" + str(rannum) + ").jpg"))

@bot.command(pass_context=True)
async def 태봄(ctx):
    rannum = random.randrange(1, 9)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Taebom/tb (" + str(rannum) + ").jpg"))

@bot.command(pass_context=True)
async def 겨울(ctx):
    rannum = random.randrange(1, 16)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Gyuerul/gu (" + str(rannum) + ").jpg"))

@bot.command(pass_context=True)
async def 쮸(ctx):
    rannum = random.randrange(1, 11)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Zzyu/zu (" + str(rannum) + ").jpg"))
    
@bot.command(pass_context=True)
async def 야옹(ctx):
    rannum = random.randrange(1, 2)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Yaong/yo (" + str(rannum) + ").jpg"))

@bot.command(pass_context=True)
async def 영서(ctx):
    rannum = random.randrange(1, 4)
    await ctx.send(file=discord.File("C:/Users/lenya/Documents/DiscordBot/pics/Youngseo/ys (" + str(rannum) + ").jpg"))